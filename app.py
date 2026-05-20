from flask import Flask, request, send_file, render_template, session, jsonify, redirect, url_for
import pandas as pd
import openpyxl
import json
import datetime
import re
import io
import os
import traceback

app = Flask(__name__)
app.secret_key = "plexure_automation_master_key"

# 建立 excel 資料夾存放路徑
EXCEL_DIR = os.path.join(os.path.dirname(__file__), 'excel')
if not os.path.exists(EXCEL_DIR):
    os.makedirs(EXCEL_DIR)

# ==========================================
# 輔助函數 (邏輯處理)
# ==========================================

def extract_main_numbers(text):
    text = str(text)
    matches = re.findall(r"(?m)^\s*(\d+)", text)
    return matches

def is_buy1get1(text):
    if not text: return False
    return "買一送一" in str(text) or bool(re.search(r"買\s*[\d一-十]+\s*送\s*[\d一-十]+", str(text)))

def clean_empty_text(val):
    if pd.isna(val): return ""
    v = str(val).strip()
    return "" if v.lower() in ["", "nan", "none", "00:00:00", "0", "0.0"] else v

def split_product_codes(r_text, promo_text):
    codes = extract_main_numbers(r_text)
    if not codes: return "", ""
    if is_buy1get1(promo_text):
        res = "|".join(codes)
        return res, res
    return "|".join(codes), (codes[-1] if codes else "")

def split_time(text):
    text = str(text).strip()
    if text.upper() == "NA" or not text or text.lower() == "nan": return "Nan", "Nan"
    if "-" in text:
        parts = text.split("-")
        return parts[0].strip(), (parts[1].strip() if len(parts) > 1 else "Nan")
    return text, "Nan"

def generate_tags(promo_text, original_tags_raw):
    tags = ["Claim Type > Scan And Go QR"]
    t = str(promo_text or "").strip()
    
    # 指向修正後的 setting.json
    config_path = os.path.join(os.path.dirname(__file__), 'static', 'json', 'setting.json')
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
            for item in config_data.get("keywords", []):
                filter_name = item.get("filter")
                keywords_list = item.get("values", [])
                if any(kw in t for kw in keywords_list):
                    tags.append(f"Deal Filter Tag > {filter_name}")
        except Exception as e:
            print(f"generate_tags 讀取 setting.json 失敗: {e}")

    orig_cleaned = str(original_tags_raw or "").strip()
    if orig_cleaned and orig_cleaned.lower() not in ["nan", "none", "0"]:
        for ot in [o.strip() for o in orig_cleaned.split(",") if o.strip()]:
            if ot not in tags:
                tags.append(ot)
    return ",".join(tags)


def get_extended_template(text):
    if pd.isna(text): return ""
    t = str(text).strip()
    
    # 指向修正後的 setting.json
    config_path = os.path.join(os.path.dirname(__file__), 'static', 'json', 'setting.json')
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
            for item in config_data.get("templates", []):
                regex_str = item.get("regex", "")
                exclude_str = item.get("exclude_regex", "")
                template_value = item.get("value", "")
                
                if regex_str and re.search(regex_str, t):
                    if exclude_str and re.search(exclude_str, t):
                        continue
                    return template_value
        except Exception as e:
            print(f"get_extended_template 讀取 setting.json 失敗: {e}")
    return ""

# ==========================================
# 核心轉換邏輯 (JSON 產出)
# ==========================================

def perform_transformation(data_path, model_path, location_path=None):
    wb_source = openpyxl.load_workbook(data_path, data_only=True)
    ws_source = wb_source.worksheets[0]
    wb_target = openpyxl.load_workbook(model_path)
    ws_target = wb_target.worksheets[0]

    # 【核心改動】如果使用者有輸入新網址，直接強制覆寫樣板的 D2 欄位 (第 2 行，第 4 欄)
    if location_path:
        ws_target.cell(row=2, column=4).value = str(location_path).strip()

    target_cells_addr = ["E9", "E19", "D20", "E24", "E26", "E28", "E30", "E32", "E34", "E36", "E38", "E40", "E42", "E47", "E52", "E54", "E56", "E58", "E60", "E62", "E64", "E66", "D69", "E71", "E73", "E75", "E77"]
    template_height = 76
    next_row = 3
    
    # 此時讀取的 d2_url 就會是使用者剛剛輸入的新網址了！
    d2_url = ws_target.cell(row=2, column=4).value

    for data_row in range(2, ws_source.max_row + 1):
        if data_row > 2:
            for r in range(template_height):
                for c in range(1, ws_target.max_column + 1):
                    ws_target.cell(row=next_row + r, column=c).value = ws_target.cell(row=3 + r, column=c).value

        for i, addr in enumerate(target_cells_addr):
            source_val = str(ws_source.cell(row=data_row, column=i+1).value or "").strip()
            orig = wb_target.worksheets[0][addr]
            tr, tc = orig.row + (next_row - 3), orig.column
            t_cell = ws_target.cell(row=tr, column=tc)
            t_cell.number_format = "@"

            if addr == "D20":
                ws_target.cell(row=tr, column=3).value = "executeScript"
                t_cell.value = f"var targetText = '{source_val}'; var $select = window.jQuery('#ExtendedDataTemplateSelector'); var $opt = $select.find('option').filter(function() {{ return window.jQuery(this).text().trim() === targetText; }}); if($opt.length > 0) {{ $select.val($opt.val()).trigger('change'); }}"
            elif addr == "D69":
                ws_target.cell(row=tr, column=3).value = "executeScript"
                t_cell.value = f"(function(){{var t='{source_val.replace("'", "\\'")}';var s=window.jQuery('#OfferDetails_CategoryId');var v=s.find('option').filter(function(){{return window.jQuery(this).text().trim().indexOf(t)>-1;}}).val();if(v){{s.val(v).trigger('change').trigger('change.select2');}}window.jQuery('.select2-result-label').filter(function(){{return window.jQuery(this).text().trim().indexOf(t)>-1;}}).click();}})();"
            elif addr == "E77":
                ws_target.cell(row=tr, column=3).value = "type"
                ws_target.cell(row=tr, column=4).value = "id=OfferDetails_TermsAndConditionsTranslated_zh_"
                t_cell.value = source_val if source_val else "每券限兌換一次。每筆交易可以同時使用多張不同品項之回饋券或優惠券。"
            else:
                t_cell.value = source_val

        # 動態追加 Tags 與 Stores
        dyn_r = next_row + template_height
        for i, col in enumerate([28, 29, 30]):
            txt = str(ws_source.cell(row=data_row, column=col).value or "").strip()
            if txt:
                for tag in [t.strip() for t in txt.split(",") if t.strip()]:
                    ws_target.cell(row=dyn_r, column=3).value = "addSelection"
                    ws_target.cell(row=dyn_r, column=4).value = f"id=allTags{i+1}"
                    ws_target.cell(row=dyn_r, column=5).value = f"label={tag}"
                    dyn_r += 1
                ws_target.cell(row=dyn_r, column=3).value = "click"
                ws_target.cell(row=dyn_r, column=4).value = f"id=btnAdd{i+1}"
                dyn_r += 1

        stores = str(ws_source.cell(row=data_row, column=31).value or "").strip()
        if stores:
            sids = [s.strip() for s in stores.split("#") if s.strip()]
            ws_target.cell(row=dyn_r, column=3).value = "click"
            ws_target.cell(row=dyn_r, column=4).value = "xpath=//input[@id='OfferStores_isAvailableAllStores' and @value='False']"
            dyn_r += 1
            ws_target.cell(row=dyn_r, column=3).value = "pause"
            ws_target.cell(row=dyn_r, column=4).value = "1500"
            dyn_r += 1
            for sid in sids:
                txp = f"xpath=//li[label[contains(text(), '{sid}')]]//input"
                ws_target.cell(row=dyn_r, column=3).value = "executeScript"
                ws_target.cell(row=dyn_r, column=4).value = f"var el = document.evaluate(\"{txp.replace('xpath=', '')}\", document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue; if(el) el.scrollIntoView();"
                dyn_r += 1
                ws_target.cell(row=dyn_r, column=3).value = "click"
                ws_target.cell(row=dyn_r, column=4).value = txp
                dyn_r += 1

        ws_target.cell(row=dyn_r, column=3).value = "click"
        ws_target.cell(row=dyn_r, column=4).value = "id=btnSave2"
        next_row = dyn_r + 1

   # --- 修正後的 JSON 產出邏輯 ---
    plexure_json = {"Name": int(datetime.datetime.now().strftime("%Y%m%d")), "CreationDate": 45951, "Commands": []}
    new_offer_xp = "xpath=//button[contains(@class, 'btn-primary') and .//span[contains(text(), 'New Offer')]]"

    # 使用 values 快速讀取整列資料
    rows = list(ws_target.iter_rows(min_row=3, max_col=5, values_only=True))
    
    for idx, row in enumerate(rows):
        curr_r = idx + 3
        cmd = str(row[2] or "").strip()    # 欄位 3
        target = str(row[3] or "").strip() # 欄位 4
        val = str(row[4] or "").strip()    # 欄位 5

        # 1. 判斷是否為每一筆資料的開頭：插入 Open 與 Click New Offer
        # 條件：行號為 3 (第一筆) 或是前一行的 Target 是 btnSave2 (代表上一筆結束)
        if curr_r == 3 or (str(ws_target.cell(row=curr_r-1, column=4).value) == "id=btnSave2"):
            plexure_json["Commands"].append({"Command": "open", "Target": str(d2_url), "Value": ""})

        # 2. 處理當前指令 (必須在迴圈內)
        if cmd or target:
            # 圖片儲存防呆捲動
            if "Promo_en_saveButton" in target or "Promo_zh_saveButton" in target:
                clean_id = target.replace('id=', '')
                plexure_json["Commands"].append({
                    "Command": "executeScript", 
                    "Target": f"document.getElementById('{clean_id}').scrollIntoView();"
                })
                plexure_json["Commands"].append({"Command": "pause", "Target": "2000"})

            # 加入指令
            new_cmd = {"Command": cmd, "Target": target}
            if val: new_cmd["Value"] = val
            plexure_json["Commands"].append(new_cmd)
            
            # 存檔後切換分頁以穩定流程
            if "btnSave2" in target:
                plexure_json["Commands"].append({"Command": "pause", "Target": "1500"})
                plexure_json["Commands"].append({"Command": "selectWindow", "Target": "tab=0"})

    return json.dumps(plexure_json, ensure_ascii=False, indent=2)

# ==========================================
# Flask 路由
# ==========================================

@app.route('/')
def index():
    if session.get('logged_in'): return redirect(url_for('step1'))
    return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    d = request.get_json()
    if d.get('username') == "admin" and d.get('password') == "12345":
        session['logged_in'] = True
        return jsonify({"status": "success"})
    return jsonify({"status": "error", "message": "帳號或密碼錯誤"})

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/step1')
def step1():
    if not session.get('logged_in'): return redirect(url_for('index'))
    return render_template('step1.html')

@app.route('/result', methods=['GET'])
def result_get():
    # 防呆：如果使用者在 /result 頁面不小心按到重新整理，自動幫他導回 step2
    return redirect(url_for('step2'))

@app.route('/transform', methods=['POST'])
def transform():
    if not session.get('logged_in'): return jsonify({"status": "error", "message": "請登入"}), 401
    try:
        file = request.files.get('file')
        image_base_path = request.form.get('imagePath', "").strip()
        location_path = request.form.get('locationPath', "").strip()

        if not file:
                    return jsonify({"status": "error", "message": "未接收到檔案"})

        # 【新增】將網址存在 Session 暫存，供 Step 2 使用
        session['location_path'] = location_path

        if image_base_path and not image_base_path.endswith(('/', '\\')): 
            image_base_path += "/"

        df = pd.read_excel(file, header=None, skiprows=2, engine='openpyxl')
        
        df = df[df.iloc[:, 11].notna()].reset_index(drop=True)
        out = pd.DataFrame()
        
        def get_img(val):
            c = clean_empty_text(val)
            return f"{image_base_path}{c}" if c else ""
        
        out["Internal Name"] = df.iloc[:, 11]
        out["Base Weight"] = df.iloc[:, 14]
        out["Extended Data Templates"] = df.iloc[:, 11].apply(get_extended_template)
        out["Promotion Name(EN)"] = df.iloc[:, 25].apply(clean_empty_text)
        out["Promotion Short Description(EN)"] = df.iloc[:, 27].apply(clean_empty_text)
        out["Promotion Long Description(EN)"] = df.iloc[:, 29].apply(clean_empty_text)
        out["Promotion Name(ZH)"] = df.iloc[:, 24].apply(clean_empty_text)
        out["Promotion Short Description(ZH)"] = df.iloc[:, 26].apply(clean_empty_text)
        out["Promotion Long Description(ZH)"] = df.iloc[:, 28].apply(clean_empty_text)
        res = df.apply(lambda r: split_product_codes(r.iloc[17], r.iloc[11]), axis=1)
        out["Product Code Buy"] = [r[0] for r in res]
        out["Product Code Discounted"] = [r[1] for r in res]
        out["Percentage"] = "1%"
        out["Promotional Image En"] = df.iloc[:, 49].apply(get_img)
        out["Promotional Image Zh"] = df.iloc[:, 50].apply(get_img)
        out["Title EN"] = df.iloc[:, 31].apply(clean_empty_text)
        out["Title CH"] = df.iloc[:, 30].apply(clean_empty_text)
        out["Start Date"] = pd.to_datetime(df.iloc[:, 9], errors="coerce").dt.strftime("%Y/%m/%d")
        out["Daily Start"] = "12:00 AM"
        out["End Date"] = pd.to_datetime(df.iloc[:, 10], errors="coerce").dt.strftime("%Y/%m/%d")
        out["Daily End"] = "11:59 PM"
        time_split = df.iloc[:, 34].apply(lambda x: pd.Series(split_time(x)))
        out["Daily Start Split"] = time_split[0]
        out["Daily End Split"] = time_split[1]
        out["Category"] = df.iloc[:, 38].apply(clean_empty_text) 
        out["Desc EN"] = df.iloc[:, 40].apply(clean_empty_text)
        out["Terms EN"] = df.iloc[:, 42].apply(clean_empty_text)
        out["Desc ZH"] = df.iloc[:, 39].apply(clean_empty_text)
        out["Terms ZH"] = df.iloc[:, 41].apply(clean_empty_text)
        out["addSelection1"] = df.iloc[:, 51].apply(clean_empty_text)
        out["addSelection2"] = ""
        # 修改後的 addSelection3 邏輯
        out["addSelection3"] = df.apply(
            lambda row: generate_tags(row.iloc[11], row.iloc[46]), 
            axis=1
        )
        out["stores"] = df.iloc[:, 47].apply(clean_empty_text)
        # 最終存到 excel/data.xlsx
        out.to_excel(os.path.join(EXCEL_DIR, 'data.xlsx'), index=False)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/step2')
def step2():
    if not session.get('logged_in'): return redirect(url_for('index'))
    models = [file for file in os.listdir(EXCEL_DIR) if file.endswith('.xlsm')]
    return render_template('step2.html', models=models)


@app.route('/setting')
def setting():
    if not session.get('logged_in'): 
        return redirect(url_for('index'))
        
    json_dir = os.path.join(os.path.dirname(__file__), 'static', 'json')
    if not os.path.exists(json_dir):
        os.makedirs(json_dir)
        
    # 【修正】精準對應你目前的單一設定檔：setting.json
    config_path = os.path.join(json_dir, 'setting.json')
    
    # 預設結構：如果檔案完全不存在時自動產生的防呆初始資料
    default_config = {
        "keywords": [
            { "filter": f"dealFilter{i}", "values": [] } for i in range(1, 10)
        ],
        "templates": [
            { "id": "tpl1", "value": "MPA TW Discount Off Product(Percentage) Pre tax False", "regex": "單點.*(?:打?\\d+折|折.*%)", "exclude_regex": "" },
            { "id": "tpl2", "value": "MPA TW Discount Off Product($Amount) Pre tax False", "regex": "單點.*折.*\\d+|買.*現折|單點.*特價", "exclude_regex": "" },
            { "id": "tpl3", "value": "TW Buy One Get One Or Another Discounted(Percentage)", "regex": "買一送一|買.*送|單點.*送|單筆.*滿.*送", "exclude_regex": "加(?:\\$|\\d+元).*送" },
            { "id": "tpl4", "value": "TW Buy One Get One Or Another Discounted($Amount)", "regex": "單點.*加(?:\\$|\\d+元).*送", "exclude_regex": "" },
            { "id": "tpl5", "value": "MPA TW Discount Off Total Order(Percentage) Pre Tax False", "regex": "單筆.*滿.*折", "exclude_regex": "" },
            { "id": "tpl6", "value": "MPA TW Discount Off Total Order($Amount) Pre tax False", "regex": "單筆.*滿.*現折.*\\$", "exclude_regex": "" }
        ]
    }

    config_data = default_config
    
    # 如果 setting.json 存在就讀取，不存在就自動寫入預設結構
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
        except Exception as e:
            print(f"讀取 setting.json 失敗，將採用預設結構: {e}")
    else:
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"自動建立 setting.json 失敗: {e}")

    # 【核心修正】建立符合前端 setting.html 命名規範的變數 'data'
    data = {f"dealFilter{i}": "" for i in range(1, 10)}
    for item in config_data.get("keywords", []):
        filter_name = item.get("filter")
        values_list = item.get("values", [])
        data[filter_name] = ",".join(values_list)

    # 取得樣板規則列表
    templates = config_data.get("templates", [])
            
    # 【核心修正】將 kw_data 改回 data 傳入前端，完美解決 jinja2 報錯
    return render_template('setting.html', data=data, templates=templates)


@app.route('/save_setting', methods=['POST'])
def save_setting():
    if not session.get('logged_in'): 
        return jsonify({"status": "error", "message": "未登入"}), 403
    try:
        req_data = request.json
        config_path = os.path.join(os.path.dirname(__file__), 'static', 'json', 'setting.json')
        
        # 先讀出目前完整的 config 物件，確保另一邊的 templates 區塊不會被洗掉
        current_config = {"keywords": [], "templates": []}
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                current_config = json.load(f)
                
        # 更新 keywords 區塊
        new_keywords = []
        for i in range(1, 10):
            filter_name = f"dealFilter{i}"
            raw_value = req_data.get(filter_name, "")
            split_values = re.split(r'[,\uff0c]', raw_value)
            clean_values = [v.strip() for v in split_values if v.strip()]
            new_keywords.append({"filter": filter_name, "values": clean_values})
            
        current_config["keywords"] = new_keywords
        
        # 寫回設定檔
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(current_config, f, ensure_ascii=False, indent=2)
            
        return jsonify({"status": "success", "message": "關鍵字已成功儲存！"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/save_template_setting', methods=['POST'])
def save_template_setting():
    if not session.get('logged_in'): 
        return jsonify({"status": "error", "message": "未登入"}), 403
    try:
        req_data = request.json  # 前端傳過來的新 templates 陣列
        config_path = os.path.join(os.path.dirname(__file__), 'static', 'json', 'setting.json')
        
        # 先讀出目前完整的 config 物件，確保另一邊的 keywords 區塊不會被洗掉
        current_config = {"keywords": [], "templates": []}
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                current_config = json.load(f)
                
        # 更新 templates 區塊
        current_config["templates"] = req_data
        
        # 寫回設定檔
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(current_config, f, ensure_ascii=False, indent=2)
            
        return jsonify({"status": "success", "message": "樣板規則已成功儲存！"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/generate_result', methods=['POST'])
def generate_result():
    try:
        selected_model = request.form.get('modelFile')
        data_p = os.path.join(EXCEL_DIR, 'data.xlsx')
        model_p = os.path.join(EXCEL_DIR, selected_model)

        # 【新增】從 Session 中取出 Step 1 填寫的網址
        location_path = session.get('location_path', "")

        # 【修改】將網址變數傳入轉換函式
        json_res = perform_transformation(data_p, model_p, location_path)

        return render_template('result.html', json_content=json_res)
    except Exception as e:
        return f"錯誤：{str(e)}", 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)